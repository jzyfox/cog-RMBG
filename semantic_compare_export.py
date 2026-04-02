from __future__ import annotations

import io
import json
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Font, PatternFill
from PIL import Image

from asset_catalog_grid import normalize_catalog_type, parse_catalog_category_from_stem
from semantic_tagger import (
    SEMANTIC_CATEGORY_SCHEMAS,
    SUPPORTED_SEMANTIC_CATEGORIES,
    SUPPORTED_SEMANTIC_CATEGORY_SET,
    normalize_semantic_record,
)

IMAGE_SUFFIXES = (".png", ".jpg", ".jpeg", ".webp")
REPORT_MIMETYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
REPORT_SHEET_NAME = "对比结果"
THUMBNAIL_SIZE = (180, 180)

COLOR_RED = "FFFF0000"
COLOR_YELLOW = "FFFFC000"
HEADER_FILL = PatternFill("solid", fgColor="D1D5DB")
ERROR_ROW_FILL = PatternFill("solid", fgColor="FEF2F2")

DISPLAY_FIELDS: list[tuple[str, str]] = [
    ("category", "品类"),
    ("primary_style", "主风格"),
    ("secondary_style", "次风格"),
    ("color_family", "基础颜色"),
    ("color_brightness", "明暗"),
    ("main_material", "主材质"),
    ("secondary_materials", "辅材质"),
    ("category_details", "品类细节"),
]


@dataclass(frozen=True)
class SidecarEntry:
    side: str
    root: Path
    json_path: Path
    image_path: Path | None
    relative_key: str
    stem: str
    filename_category: str


@dataclass(frozen=True)
class ReportRow:
    absolute_path: str
    file_name: str
    preview_image_path: Path | None
    online_display: str | CellRichText
    local_display: str | CellRichText
    difference_display: str | CellRichText
    is_exception: bool
    sort_key: tuple[int, str, str]


def export_semantic_compare_report(
    online_dir: str | Path,
    local_dir: str | Path,
    category: str | Path,
) -> tuple[io.BytesIO, str]:
    online_root = _resolve_input_dir(online_dir, "在线模型标签目录")
    local_root = _resolve_input_dir(local_dir, "本地 8B 标签目录")
    selected_category = _normalize_selected_category(category)

    rows = _build_report_rows(
        online_root=online_root,
        local_root=local_root,
        selected_category=selected_category,
    )

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = REPORT_SHEET_NAME

    image_refs: list[io.BytesIO] = []
    _build_report_sheet(sheet, rows, image_refs)

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return output, f"semantic_compare_{selected_category}_{timestamp}.xlsx"


def _resolve_input_dir(raw_path: str | Path, label: str) -> Path:
    path = Path(str(raw_path or "")).expanduser().resolve()
    if not str(raw_path or "").strip():
        raise ValueError(f"请填写{label}")
    if not path.exists():
        raise FileNotFoundError(f"{label}不存在：{path}")
    if not path.is_dir():
        raise NotADirectoryError(f"{label}不是文件夹：{path}")
    return path


def _normalize_selected_category(raw_category: str | Path) -> str:
    normalized = normalize_catalog_type(str(raw_category or "").strip().lower())
    if normalized not in SUPPORTED_SEMANTIC_CATEGORY_SET:
        supported = ", ".join(SUPPORTED_SEMANTIC_CATEGORIES)
        raise ValueError(f"品类必须是以下语义支持品类之一：{supported}")
    return normalized


def _build_report_rows(
    online_root: Path,
    local_root: Path,
    selected_category: str,
) -> list[ReportRow]:
    online_entries = _scan_sidecar_entries(online_root, side="在线", selected_category=selected_category)
    local_entries = _scan_sidecar_entries(local_root, side="本地", selected_category=selected_category)
    paired_entries, exceptions = _pair_sidecar_entries(
        online_entries=online_entries,
        local_entries=local_entries,
        selected_category=selected_category,
    )

    rows: list[ReportRow] = []
    rows.extend(_build_exception_rows(exceptions))

    for online_entry, local_entry in paired_entries:
        pair_errors = _validate_pair_entries(
            online_entry=online_entry,
            local_entry=local_entry,
            selected_category=selected_category,
        )
        if pair_errors:
            rows.extend(_build_exception_rows(pair_errors))
            continue

        online_record, online_error, online_declared_category, online_error_type = _load_normalized_record(online_entry)
        local_record, local_error, local_declared_category, local_error_type = _load_normalized_record(local_entry)
        if online_error or local_error:
            if online_error_type == "category_mismatch" or local_error_type == "category_mismatch":
                rows.append(
                    _build_exception_report_row(
                        error_type="category_mismatch",
                        selected_category=selected_category,
                        online_entry=online_entry,
                        local_entry=local_entry,
                        message="sidecar 声明的 category 与所选品类或图片文件名后缀不一致",
                        online_category=online_declared_category,
                        local_category=local_declared_category,
                    )
                )
            else:
                if online_error:
                    rows.append(
                        _build_exception_report_row(
                            error_type="json_error",
                            selected_category=selected_category,
                            online_entry=online_entry,
                            local_entry=local_entry,
                            message=online_error,
                            online_category=online_declared_category,
                        )
                    )
                if local_error:
                    rows.append(
                        _build_exception_report_row(
                            error_type="json_error",
                            selected_category=selected_category,
                            online_entry=online_entry,
                            local_entry=local_entry,
                            message=local_error,
                            local_category=local_declared_category,
                        )
                    )
            continue

        assert online_record is not None
        assert local_record is not None

        online_actual_category = str(online_record.get("category") or "").strip()
        local_actual_category = str(local_record.get("category") or "").strip()
        if online_actual_category != selected_category or local_actual_category != selected_category:
            rows.append(
                _build_exception_report_row(
                    error_type="category_mismatch",
                    selected_category=selected_category,
                    online_entry=online_entry,
                    local_entry=local_entry,
                    message="文件名后缀命中所选品类，但规范化后的 sidecar category 与所选品类不一致",
                    online_category=online_actual_category,
                    local_category=local_actual_category,
                )
            )
            continue

        rows.append(_build_normal_report_row(online_entry, local_entry, online_record, local_record))

    rows.sort(key=lambda item: item.sort_key)
    return rows


def _scan_sidecar_entries(root: Path, side: str, selected_category: str) -> list[SidecarEntry]:
    entries: list[SidecarEntry] = []
    for json_path in sorted(path for path in root.rglob("*.json") if path.is_file()):
        image_path = _find_neighbor_image(json_path)
        filename_category = _parse_filename_category(image_path.stem if image_path else json_path.stem)
        if filename_category != selected_category:
            continue
        entries.append(
            SidecarEntry(
                side=side,
                root=root,
                json_path=json_path,
                image_path=image_path,
                relative_key=json_path.relative_to(root).with_suffix("").as_posix(),
                stem=json_path.stem,
                filename_category=filename_category,
            )
        )
    return entries


def _find_neighbor_image(json_path: Path) -> Path | None:
    for suffix in IMAGE_SUFFIXES:
        candidate = json_path.with_suffix(suffix)
        if candidate.exists() and candidate.is_file():
            return candidate
    return None


def _parse_filename_category(stem: str) -> str:
    return parse_catalog_category_from_stem(stem)


def _pair_sidecar_entries(
    online_entries: list[SidecarEntry],
    local_entries: list[SidecarEntry],
    selected_category: str,
) -> tuple[list[tuple[SidecarEntry, SidecarEntry]], list[dict[str, str]]]:
    paired: list[tuple[SidecarEntry, SidecarEntry]] = []
    exceptions: list[dict[str, str]] = []

    online_by_rel = {entry.relative_key: entry for entry in online_entries}
    local_by_rel = {entry.relative_key: entry for entry in local_entries}

    matched_online: set[Path] = set()
    matched_local: set[Path] = set()

    for relative_key in sorted(set(online_by_rel) & set(local_by_rel)):
        online_entry = online_by_rel[relative_key]
        local_entry = local_by_rel[relative_key]
        paired.append((online_entry, local_entry))
        matched_online.add(online_entry.json_path)
        matched_local.add(local_entry.json_path)

    remaining_online = [entry for entry in online_entries if entry.json_path not in matched_online]
    remaining_local = [entry for entry in local_entries if entry.json_path not in matched_local]

    online_by_stem = _group_entries_by_stem(remaining_online)
    local_by_stem = _group_entries_by_stem(remaining_local)

    for stem in sorted(set(online_by_stem) | set(local_by_stem)):
        online_group = online_by_stem.get(stem, [])
        local_group = local_by_stem.get(stem, [])

        if len(online_group) == 1 and len(local_group) == 1:
            paired.append((online_group[0], local_group[0]))
            continue

        if online_group and not local_group:
            for entry in online_group:
                exceptions.append(
                    _build_exception_meta(
                        error_type="missing_counterpart",
                        selected_category=selected_category,
                        online_entry=entry,
                        local_entry=None,
                        message="在线目录存在目标品类 sidecar，但本地目录未找到可匹配项",
                    )
                )
            continue

        if local_group and not online_group:
            for entry in local_group:
                exceptions.append(
                    _build_exception_meta(
                        error_type="missing_counterpart",
                        selected_category=selected_category,
                        online_entry=None,
                        local_entry=entry,
                        message="本地目录存在目标品类 sidecar，但在线目录未找到可匹配项",
                    )
                )
            continue

        if len(online_group) > 1:
            exceptions.append(
                _build_exception_meta(
                    error_type="duplicate_stem",
                    selected_category=selected_category,
                    online_entry=None,
                    local_entry=None,
                    message=f"同名 stem 在在线目录中不唯一，无法仅按文件名回退配对（{len(online_group)} 条）",
                    stem=stem,
                    online_relative_path=",".join(item.relative_key for item in online_group),
                    local_relative_path=",".join(item.relative_key for item in local_group),
                )
            )
        if len(local_group) > 1:
            exceptions.append(
                _build_exception_meta(
                    error_type="duplicate_stem",
                    selected_category=selected_category,
                    online_entry=None,
                    local_entry=None,
                    message=f"同名 stem 在本地目录中不唯一，无法仅按文件名回退配对（{len(local_group)} 条）",
                    stem=stem,
                    online_relative_path=",".join(item.relative_key for item in online_group),
                    local_relative_path=",".join(item.relative_key for item in local_group),
                )
            )

    return paired, exceptions


def _group_entries_by_stem(entries: list[SidecarEntry]) -> dict[str, list[SidecarEntry]]:
    grouped: dict[str, list[SidecarEntry]] = {}
    for entry in entries:
        grouped.setdefault(entry.stem, []).append(entry)
    return grouped


def _validate_pair_entries(
    online_entry: SidecarEntry,
    local_entry: SidecarEntry,
    selected_category: str,
) -> list[dict[str, str]]:
    errors: list[dict[str, str]] = []
    if online_entry.image_path is None:
        errors.append(
            _build_exception_meta(
                error_type="image_missing",
                selected_category=selected_category,
                online_entry=online_entry,
                local_entry=local_entry,
                message="在线 sidecar 缺少同名图片",
            )
        )
    if local_entry.image_path is None:
        errors.append(
            _build_exception_meta(
                error_type="image_missing",
                selected_category=selected_category,
                online_entry=online_entry,
                local_entry=local_entry,
                message="本地 sidecar 缺少同名图片",
            )
        )
    return errors


def _load_normalized_record(entry: SidecarEntry) -> tuple[dict[str, Any] | None, str | None, str, str]:
    if entry.image_path is None:
        return None, f"{entry.side} sidecar 缺少同名图片", "", "image_missing"

    try:
        raw = json.loads(entry.json_path.read_text(encoding="utf-8"))
    except Exception as exc:
        return None, f"{entry.side} JSON 解析失败：{exc}", "", "json_error"

    declared_category = _extract_declared_category(raw)
    if declared_category and declared_category != entry.filename_category:
        return (
            None,
            f"{entry.side} sidecar 声明的 category 与图片文件名后缀不一致：{declared_category}",
            declared_category,
            "category_mismatch",
        )

    try:
        normalized = normalize_semantic_record(
            raw,
            image_path=entry.image_path,
            input_root=entry.root,
            vlm_model=raw.get("vlm_model") if isinstance(raw, dict) else None,
            prompt_version=raw.get("prompt_version") if isinstance(raw, dict) else None,
        )
    except Exception as exc:
        return None, f"{entry.side} JSON 校验失败：{exc}", declared_category, "json_error"
    return normalized, None, declared_category or str(normalized.get("category") or "").strip(), ""


def _extract_declared_category(raw: Any) -> str:
    if not isinstance(raw, dict):
        return ""
    raw_category = str(raw.get("category") or "").strip().lower()
    if not raw_category:
        return ""
    return normalize_catalog_type(raw_category)


def _build_normal_report_row(
    online_entry: SidecarEntry,
    local_entry: SidecarEntry,
    online_record: dict[str, Any],
    local_record: dict[str, Any],
) -> ReportRow:
    field_results = _build_field_results(online_record, local_record)
    preview_path = _choose_preview_path(online_entry, local_entry)
    absolute_path = str((preview_path or online_entry.json_path or local_entry.json_path).resolve())
    file_name = (preview_path or online_entry.json_path or local_entry.json_path).name
    sort_path = _build_display_relative_path(online_entry, local_entry)

    return ReportRow(
        absolute_path=absolute_path,
        file_name=file_name,
        preview_image_path=preview_path,
        online_display=_build_side_rich_text(field_results, side="online"),
        local_display=_build_side_rich_text(field_results, side="local"),
        difference_display=_build_difference_rich_text(field_results),
        is_exception=False,
        sort_key=(0, sort_path, file_name),
    )


def _build_exception_rows(exceptions: list[dict[str, str]]) -> list[ReportRow]:
    return [_build_exception_report_row_from_meta(item) for item in exceptions]


def _build_exception_report_row_from_meta(exception_meta: dict[str, str]) -> ReportRow:
    online_path = exception_meta.get("online_absolute_path", "")
    local_path = exception_meta.get("local_absolute_path", "")
    preview_path_str = exception_meta.get("preview_image_path", "")
    preview_path = Path(preview_path_str) if preview_path_str else None
    file_name = exception_meta.get("file_name", "") or Path(online_path or local_path or exception_meta.get("stem", "")).name

    return ReportRow(
        absolute_path=online_path or local_path,
        file_name=file_name,
        preview_image_path=preview_path,
        online_display=str(exception_meta.get("online_display", "")),
        local_display=str(exception_meta.get("local_display", "")),
        difference_display=_make_rich_text(
            [
                (f"异常: {exception_meta.get('type', '')}", COLOR_RED),
                ("\n", None),
                (str(exception_meta.get("message", "")), COLOR_RED),
            ]
        ),
        is_exception=True,
        sort_key=(1, str(exception_meta.get("online_relative_path", "") or exception_meta.get("local_relative_path", "")), file_name),
    )


def _build_exception_report_row(
    *,
    error_type: str,
    selected_category: str,
    online_entry: SidecarEntry | None,
    local_entry: SidecarEntry | None,
    message: str,
    online_category: str = "",
    local_category: str = "",
) -> ReportRow:
    return _build_exception_report_row_from_meta(
        _build_exception_meta(
            error_type=error_type,
            selected_category=selected_category,
            online_entry=online_entry,
            local_entry=local_entry,
            message=message,
            online_category=online_category,
            local_category=local_category,
        )
    )


def _build_exception_meta(
    *,
    error_type: str,
    selected_category: str,
    online_entry: SidecarEntry | None,
    local_entry: SidecarEntry | None,
    message: str,
    online_category: str = "",
    local_category: str = "",
    stem: str = "",
    online_relative_path: str = "",
    local_relative_path: str = "",
) -> dict[str, str]:
    preview_path = _choose_preview_path(online_entry, local_entry)
    online_source_path = _choose_source_path(online_entry)
    local_source_path = _choose_source_path(local_entry)
    file_name = (preview_path or online_source_path or local_source_path or Path(stem)).name if (preview_path or online_source_path or local_source_path or stem) else ""

    online_lines: list[str] = []
    if online_entry or online_relative_path:
        online_lines.append(f"路径: {online_relative_path or (online_entry.relative_key if online_entry else '')}")
        if online_category:
            online_lines.append(f"实际品类: {online_category}")
    else:
        online_lines.append("无在线记录")

    local_lines: list[str] = []
    if local_entry or local_relative_path:
        local_lines.append(f"路径: {local_relative_path or (local_entry.relative_key if local_entry else '')}")
        if local_category:
            local_lines.append(f"实际品类: {local_category}")
    else:
        local_lines.append("无本地记录")

    return {
        "type": error_type,
        "selected_category": selected_category,
        "online_relative_path": online_relative_path or (online_entry.relative_key if online_entry else ""),
        "local_relative_path": local_relative_path or (local_entry.relative_key if local_entry else ""),
        "online_category": online_category,
        "local_category": local_category,
        "stem": stem or ((online_entry or local_entry).stem if (online_entry or local_entry) else ""),
        "message": message,
        "online_absolute_path": str(online_source_path.resolve()) if online_source_path else "",
        "local_absolute_path": str(local_source_path.resolve()) if local_source_path else "",
        "preview_image_path": str(preview_path.resolve()) if preview_path else "",
        "file_name": file_name,
        "online_display": "\n".join(online_lines),
        "local_display": "\n".join(local_lines),
    }


def _choose_preview_path(online_entry: SidecarEntry | None, local_entry: SidecarEntry | None) -> Path | None:
    return (online_entry.image_path if online_entry else None) or (local_entry.image_path if local_entry else None)


def _choose_source_path(entry: SidecarEntry | None) -> Path | None:
    if entry is None:
        return None
    return entry.image_path or entry.json_path


def _build_display_relative_path(online_entry: SidecarEntry, local_entry: SidecarEntry) -> str:
    if online_entry.relative_key == local_entry.relative_key:
        return online_entry.relative_key
    return f"{online_entry.relative_key} <> {local_entry.relative_key}"


def _build_field_results(online_record: dict[str, Any], local_record: dict[str, Any]) -> list[dict[str, str]]:
    results: list[dict[str, str]] = []
    for field_name, field_label in DISPLAY_FIELDS:
        online_value = _extract_display_value(online_record, field_name)
        local_value = _extract_display_value(local_record, field_name)
        results.append(
            {
                "field_name": field_name,
                "field_label": field_label,
                "online_value": online_value,
                "local_value": local_value,
                "status": _compare_display_values(online_value, local_value),
            }
        )
    return results


def _extract_display_value(record: dict[str, Any], field_name: str) -> str:
    if field_name == "secondary_materials":
        values = record.get("secondary_materials", [])
        if not isinstance(values, list):
            return ""
        return "、".join(str(value).strip() for value in values if str(value).strip())

    if field_name == "category_details":
        details = record.get("category_details", {})
        if not isinstance(details, dict):
            return ""
        category = str(record.get("category") or "").strip()
        schema = SEMANTIC_CATEGORY_SCHEMAS.get(category, {})
        detail_fields = schema.get("detail_fields", []) if isinstance(schema, dict) else []
        labels = {
            str(field.get("name") or ""): str(field.get("label") or field.get("name") or "")
            for field in detail_fields
            if isinstance(field, dict)
        }
        ordered_keys = [
            str(field.get("name") or "")
            for field in detail_fields
            if isinstance(field, dict) and str(field.get("name") or "") in details
        ]
        ordered_keys.extend(sorted(key for key in details.keys() if key not in labels))

        parts: list[str] = []
        seen: set[str] = set()
        for key in ordered_keys:
            if key in seen:
                continue
            seen.add(key)
            value = str(details.get(key) or "").strip()
            if not value:
                continue
            parts.append(f"{labels.get(key, key)}: {value}")
        return " / ".join(parts)

    value = record.get(field_name)
    if value in (None, ""):
        return ""
    return str(value).strip()


def _compare_display_values(online_value: str, local_value: str) -> str:
    if not online_value and not local_value:
        return "both_empty"
    if online_value == local_value:
        return "same"
    if online_value and not local_value:
        return "local_missing"
    if local_value and not online_value:
        return "online_missing"
    return "different"


def _build_side_rich_text(field_results: list[dict[str, str]], side: str) -> CellRichText:
    segments: list[tuple[str, str | None]] = []
    for index, item in enumerate(field_results):
        if index > 0:
            segments.append(("\n", None))
        color = _get_side_color(item["status"], side)
        line = f'{item["field_label"]}: {item["online_value"] if side == "online" else item["local_value"] or "-"}'
        if side == "online" and not item["online_value"]:
            line = f'{item["field_label"]}: -'
        if side == "local" and not item["local_value"]:
            line = f'{item["field_label"]}: -'
        segments.append((line, color))
    return _make_rich_text(segments)


def _get_side_color(status: str, side: str) -> str | None:
    if status == "different":
        return COLOR_YELLOW
    if status == "online_missing" and side == "online":
        return COLOR_RED
    if status == "local_missing" and side == "local":
        return COLOR_RED
    return None


def _build_difference_rich_text(field_results: list[dict[str, str]]) -> str | CellRichText:
    missing_fields = [
        item["field_label"]
        for item in field_results
        if item["status"] in {"local_missing", "online_missing"}
    ]
    different_fields = [
        item["field_label"]
        for item in field_results
        if item["status"] == "different"
    ]

    if not missing_fields and not different_fields:
        return "一致"

    segments: list[tuple[str, str | None]] = []
    if missing_fields:
        segments.append(("缺失：", None))
        segments.extend(_join_label_segments(missing_fields, COLOR_RED))
    if different_fields:
        if segments:
            segments.append(("\n", None))
        segments.append(("不同：", None))
        segments.extend(_join_label_segments(different_fields, COLOR_YELLOW))
    return _make_rich_text(segments)


def _join_label_segments(labels: list[str], color: str) -> list[tuple[str, str | None]]:
    segments: list[tuple[str, str | None]] = []
    for index, label in enumerate(labels):
        if index > 0:
            segments.append(("、", None))
        segments.append((label, color))
    return segments


def _make_rich_text(segments: list[tuple[str, str | None]]) -> CellRichText:
    rich = CellRichText()
    for text, color in segments:
        if not text:
            continue
        if color:
            rich.append(TextBlock(InlineFont(color=color), text))
        else:
            rich.append(text)
    return rich


def _build_report_sheet(sheet, rows: list[ReportRow], image_refs: list[io.BytesIO]) -> None:
    headers = ["绝对路径", "文件名", "图片", "在线模型", "本地", "差异摘要"]
    sheet.append(headers)
    _style_header_row(sheet, 1)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:F{max(len(rows) + 1, 1)}"
    sheet.sheet_view.showGridLines = False
    sheet.column_dimensions["A"].width = 52
    sheet.column_dimensions["B"].width = 24
    sheet.column_dimensions["C"].width = 18
    sheet.column_dimensions["D"].width = 38
    sheet.column_dimensions["E"].width = 38
    sheet.column_dimensions["F"].width = 28

    if not rows:
        sheet.append(["", "", "", "没有找到可导出的记录", "", ""])
        _style_data_row(sheet, 2)
        return

    for index, row in enumerate(rows, start=2):
        sheet.cell(row=index, column=1, value=row.absolute_path)
        sheet.cell(row=index, column=2, value=row.file_name)
        sheet.cell(row=index, column=4, value=row.online_display)
        sheet.cell(row=index, column=5, value=row.local_display)
        sheet.cell(row=index, column=6, value=row.difference_display)
        _style_data_row(sheet, index)
        sheet.row_dimensions[index].height = 108

        if row.is_exception:
            for col in ("A", "B", "D", "E", "F"):
                sheet[f"{col}{index}"].fill = ERROR_ROW_FILL

        image = _build_excel_thumbnail(row.preview_image_path)
        if image is not None:
            excel_image, buffer = image
            image_refs.append(buffer)
            sheet.add_image(excel_image, f"C{index}")


def _style_header_row(sheet, row_index: int) -> None:
    for cell in sheet[row_index]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = HEADER_FILL


def _style_data_row(sheet, row_index: int) -> None:
    for col in ("A", "B", "D", "E", "F"):
        sheet[f"{col}{row_index}"].alignment = Alignment(vertical="top", wrap_text=True)
    sheet[f"C{row_index}"].alignment = Alignment(horizontal="center", vertical="center")


def _build_excel_thumbnail(image_path: Path | None) -> tuple[ExcelImage, io.BytesIO] | None:
    if image_path is None:
        return None
    try:
        with Image.open(image_path) as source_image:
            rendered = source_image.convert("RGBA")
            rendered.thumbnail(THUMBNAIL_SIZE)

            background = Image.new("RGBA", THUMBNAIL_SIZE, (255, 255, 255, 255))
            offset_x = max((THUMBNAIL_SIZE[0] - rendered.width) // 2, 0)
            offset_y = max((THUMBNAIL_SIZE[1] - rendered.height) // 2, 0)
            background.paste(rendered, (offset_x, offset_y), rendered)

            buffer = io.BytesIO()
            background.convert("RGB").save(buffer, format="PNG")
            buffer.seek(0)
            buffer.name = f"{image_path.stem}.png"
            excel_image = ExcelImage(buffer)
            excel_image.width = 92
            excel_image.height = 92
            return excel_image, buffer
    except Exception:
        return None
